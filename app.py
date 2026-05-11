"""
Intrastat PDF Extractor  —  NBB/Intrastat Excel + CSV generator
Auteur: Lander Smits

PDF-verwerking via PyMuPDF (fitz) — veel lichter qua geheugen dan pdfplumber.
Beveiliging: APP_PASSWORD env-var + Flask-sessies.
"""
import csv
import gc
import io
import os
import re
import time
import uuid
import secrets
import logging
import threading
from collections import Counter, OrderedDict
from datetime import datetime
from functools import wraps

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import (Flask, request, jsonify, send_file, render_template,
                   abort, session, redirect, url_for)
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_STORE = {}
STORE_LOCK   = threading.Lock()
APP_PASSWORD = os.environ.get("APP_PASSWORD", "")

# ---------------------------------------------------------------------------
# Auto-opruiming
# ---------------------------------------------------------------------------

def _cleanup_loop():
    while True:
        time.sleep(300)
        cutoff = time.time() - 600
        with STORE_LOCK:
            old = [k for k, v in RESULT_STORE.items() if v.get("ts", 0) < cutoff]
            for k in old:
                del RESULT_STORE[k]
        if old:
            logger.info("Opgeruimd: %d verlopen resultaten", len(old))

threading.Thread(target=_cleanup_loop, daemon=True).start()

# ---------------------------------------------------------------------------
# Beveiliging
# ---------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if APP_PASSWORD and not session.get("authenticated"):
            if request.method == "POST" or request.path.startswith("/download"):
                return jsonify({"error": "Sessie verlopen. Herlaad de pagina."}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


@app.route("/login", methods=["GET", "POST"])
def login():
    if not APP_PASSWORD:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        pw = request.form.get("password", "")
        if secrets.compare_digest(pw, APP_PASSWORD):
            session["authenticated"] = True
            return redirect(url_for("index"))
        error = "Ongeldig wachtwoord"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ---------------------------------------------------------------------------
# Constanten
# ---------------------------------------------------------------------------

FIXED_GEWEST   = "1 Vlaams gewest"
FIXED_LAND     = "Italie"
FIXED_AARD     = ("Rechtstreekse verkoop/aankoop, behalve rechtstreekse handel "
                  "met/door particuliere consumenten")
FIXED_INCOTERM = "EXW Af fabriek"
FIXED_VERVOER  = "3 Wegvervoer"
NO_SUPPL_PREFIXES = ("42",)

TOTAL_PATTERNS = [
    r"Non\s+imp\.?\s+Art\s+\d+\s+\S+\s+\S+\s+([\d.,]+)",
    r"EUR\s+([\d]{2,3}[.,]\d{3}[.,]\d{2})",
    r"TOTAL\s+FREE\s+([\d.,]+)",
    r"\bVAT\b\s+([\d]{2,3}[.,]\d{3}[.,]\d{2})",
]

# ---------------------------------------------------------------------------
# Regex
# ---------------------------------------------------------------------------

MADE_IN_RE = re.compile(r"Made\s+In\s+([A-Z]{2})", re.IGNORECASE)
WEIGHT_RE  = re.compile(r"[Ww]eight\s+gr\s+(\d+)")
CN8_RE     = re.compile(r"\b(\d{8})\b")
SKIP_RE    = re.compile(
    r"(^total|transport|^iban|^bank|^date|^datum|^payment|^paym|"
    r"^delivery|appearance|signature|pallet|track no|corriere|"
    r"taxable|exempt|tva|^vat|intracommunautair)",
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_num(val):
    if val is None:
        return None
    v = str(val).strip().replace("\xa0", "").replace(" ", "")
    v = re.sub(r"[€$£]", "", v)
    if not v:
        return None
    if re.search(r"\d\.\d{3},", v):
        v = v.replace(".", "").replace(",", ".")
    elif v.count(",") == 1 and v.count(".") == 0:
        v = v.replace(",", ".")
    elif v.count(",") >= 1 and v.count(".") == 1:
        v = v.replace(",", "")
    elif v.count(".") == 1 and len(v.split(".")[1]) == 3:
        v = v.replace(".", "")
    try:
        return float(v)
    except ValueError:
        return None


def has_suppl_unit(cn8):
    return not cn8.startswith(NO_SUPPL_PREFIXES)


def _update_meta(meta, text):
    if not meta["factuurnummer"]:
        for pat in [
            r"(?:factuur|invoice|fattura)\s*(?:nr|number|no|n)?\.?\s*:?\s*([A-Z0-9\-/_]{3,30})",
            r"CFAK-(\d+[A-Za-z0-9\-]*)",
            r"CORRISPETTIVO\s*\n?\s*(\d{6,12})",
        ]:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                meta["factuurnummer"] = m.group(1).strip()
                break
    if not meta["datum"]:
        for pat in [
            r"(?:factuurdatum|datum|date)[:\s]+(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})",
            r"(\d{2}/\d{2}/\d{4})",
        ]:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                meta["datum"] = m.group(1)
                break
    if not meta["leverancier"]:
        for pat in [r"leverancier\s*:\s*(.+)", r"supplier\s*:\s*(.+)"]:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                meta["leverancier"] = m.group(1).strip()
                break

# ---------------------------------------------------------------------------
# PDF parser — PyMuPDF, pagina-voor-pagina
# ---------------------------------------------------------------------------

def _process_table(raw_table, lines, pn):
    """Verwerk een ruwe tabel (list of list of str) naar productlijnen."""
    if not raw_table or len(raw_table) < 2:
        return

    # Zoek headerrij
    hkw  = re.compile(r"(quantity|price|discount|value)", re.I)
    hidx = 0
    for i, row in enumerate(raw_table[:5]):
        if sum(1 for c in (row or []) if c and hkw.search(str(c))) >= 3:
            hidx = i
            break

    hdrs = [str(c).lower().strip() if c else "" for c in raw_table[hidx]]

    has_val = any(
        "value" in h and "taxable" not in h and "free" not in h
        for h in hdrs
    )
    if not has_val:
        return

    def find_col(patterns, exclude=None):
        for i, h in enumerate(hdrs):
            if exclude and any(e in h for e in exclude):
                continue
            if any(p in h for p in patterns):
                return i
        return None

    i_cn  = find_col(["customs"])
    i_qty = find_col(["quantity"])
    i_val = find_col(["value"], exclude=["taxable", "free"])

    if i_val is None:
        return

    for row in raw_table[hidx + 1:]:
        if not row or not any(row):
            continue

        def get(idx):
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx]).strip() if row[idx] is not None else ""

        m = CN8_RE.search(get(i_cn))
        if not m:
            continue
        cn8 = m.group(1)

        if SKIP_RE.search(" ".join(str(c) for c in row if c)):
            continue

        value = clean_num(get(i_val))
        if value is None or value <= 0:
            continue

        qty_raw = clean_num(get(i_qty))
        qty     = int(qty_raw) if qty_raw is not None else 0

        art_cell  = get(0)
        wm        = WEIGHT_RE.search(art_cell)
        weight_kg = round(int(wm.group(1)) / 1000, 4) if wm else None

        mm      = MADE_IN_RE.search(art_cell)
        made_in = mm.group(1).upper() if mm else ""

        lines.append({
            "pn": pn, "cn8": cn8, "qty": qty,
            "value": value, "weight_kg": weight_kg, "made_in": made_in,
        })


def parse_pdf(pdf_bytes):
    """
    Verwerkt PDF via PyMuPDF (fitz).
    Veel lichter qua geheugen dan pdfplumber.
    Retourneert (lines, pdf_total, meta).
    """
    lines            = []
    meta             = {"factuurnummer": "", "datum": "", "leverancier": ""}
    total_candidates = []

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    try:
        for pn in range(len(doc)):
            page      = doc[pn]
            page_text = page.get_text()

            # Meta en totaal uit tekst
            _update_meta(meta, page_text)
            for pat in TOTAL_PATTERNS:
                for m in re.finditer(pat, page_text, re.IGNORECASE | re.MULTILINE):
                    v = clean_num(m.group(1))
                    if v and v > 100:
                        total_candidates.append(round(v, 2))

            # Tabelextractie via PyMuPDF
            try:
                finder = page.find_tables()
                for tab in finder.tables:
                    raw = tab.extract()
                    _process_table(raw, lines, pn + 1)
            except Exception as te:
                logger.warning("Tabel pagina %d overgeslagen: %s", pn + 1, te)

            # Geheugen vrijgeven
            page = None
            gc.collect()

    finally:
        doc.close()
        del pdf_bytes
        gc.collect()

    pdf_total = (Counter(total_candidates).most_common(1)[0][0]
                 if total_candidates else None)

    logger.info("Geextraheerd: %d productlijnen, totaal EUR %.2f",
                len(lines), sum(r["value"] for r in lines))
    return lines, pdf_total, meta


def aggregate_by_cn8(lines):
    groups = OrderedDict()
    for line in lines:
        cn8 = line["cn8"]
        if cn8 not in groups:
            groups[cn8] = {"cn8": cn8, "qty": 0, "value": 0.0, "weight_kg": 0.0}
        groups[cn8]["qty"]      += line["qty"]
        groups[cn8]["value"]     = round(groups[cn8]["value"] + line["value"], 2)
        groups[cn8]["weight_kg"] = round(
            groups[cn8]["weight_kg"] + (line["weight_kg"] or 0), 4)
    return list(groups.values())

# ---------------------------------------------------------------------------
# NBB-rij
# ---------------------------------------------------------------------------

def build_nbb_row(agg):
    cn8 = agg["cn8"]
    return {
        "goederencode":  cn8,
        "gewest":        FIXED_GEWEST,
        "land":          FIXED_LAND,
        "aard":          FIXED_AARD,
        "incoterm":      FIXED_INCOTERM,
        "vervoer":       FIXED_VERVOER,
        "netto_gewicht": agg["weight_kg"] if agg["weight_kg"] else 0,
        "waarde":        agg["value"],
        "waarde_eur":    agg["value"],
        "aanv_eenheden": agg["qty"] if has_suppl_unit(cn8) else 0,
        "aanv_eenh":     "p/st" if has_suppl_unit(cn8) else None,
    }

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

HEADERS    = ["Goederencode", "Gewest", "Land", "Aard transactie",
              "Incoterm", "Vervoerswijze", "Netto gewicht",
              "Waarde", "Waarde EUR", "Aanvullende eenheden", "Aanv. Eenh."]
COL_WIDTHS = [14, 18, 10, 65, 15, 15, 14, 12, 12, 22, 12]
BROWN      = "8B7355"
BEIGE_ALT  = "F5F0E8"
BEIGE_WHT  = "FEFCF9"
GOLD       = "C9A96E"


def _side(color="D4C5A9"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(nbb_rows, meta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Intrastat"
    ncols = len(HEADERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value="Intrastat PDF Extractor  —  NBB Intrastat")
    t.font      = Font(bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=BROWN)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    parts = []
    if meta.get("leverancier"):   parts.append("Leverancier: " + meta["leverancier"])
    if meta.get("factuurnummer"): parts.append("Factuurnr: "   + meta["factuurnummer"])
    if meta.get("datum"):         parts.append("Datum: "       + meta["datum"])

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1, value="   ".join(parts))
    ws.cell(row=2, column=1).font      = Font(italic=True, size=9, color="5C4A32")
    ws.cell(row=2, column=1).fill      = PatternFill("solid", fgColor="EDE5D0")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", indent=1, vertical="center")

    ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=ncols)
    ws.cell(row=2, column=7, value="Export: " + datetime.today().strftime("%d/%m/%Y %H:%M"))
    ws.cell(row=2, column=7).font      = Font(italic=True, size=9, color="5C4A32")
    ws.cell(row=2, column=7).fill      = PatternFill("solid", fgColor="EDE5D0")
    ws.cell(row=2, column=7).alignment = Alignment(horizontal="right", indent=1, vertical="center")
    ws.row_dimensions[2].height = 14

    ws.append([])

    HR = 4
    for ci, (label, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=HR, column=ci, value=label)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=BROWN)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(ci == 4))
        c.border    = _side("FFFFFF")
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HR].height = 24

    DS = HR + 1
    fmt_kg = "#,##0.000"
    fmt_eu = "#,##0.00"
    fmt_qt = "#,##0"

    for ri, row in enumerate(nbb_rows):
        r    = DS + ri
        fill = PatternFill("solid", fgColor=BEIGE_ALT if ri % 2 == 0 else BEIGE_WHT)
        brd  = _side()
        vals = [row["goederencode"], row["gewest"], row["land"], row["aard"],
                row["incoterm"], row["vervoer"],
                row["netto_gewicht"], row["waarde"], row["waarde_eur"],
                row["aanv_eenheden"], row["aanv_eenh"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = fill; c.border = brd
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if ci in (7, 8, 9, 10) else "left",
                                    wrap_text=(ci == 4))
            if ci == 7 and isinstance(v, (int, float)):       c.number_format = fmt_kg
            if ci in (8, 9) and isinstance(v, (int, float)):  c.number_format = fmt_eu
            if ci == 10 and isinstance(v, (int, float)):      c.number_format = fmt_qt
        ws.row_dimensions[r].height = 15

    TR  = DS + len(nbb_rows)
    tf  = PatternFill("solid", fgColor=GOLD)
    tft = Font(bold=True, size=10)
    for ci in range(1, ncols + 1):
        c = ws.cell(row=TR, column=ci)
        c.fill = tf; c.border = _side(); c.font = tft
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=TR, column=1, value="TOTAAL").font = tft
    ws.cell(row=TR, column=1).fill = tf
    for ci, fmt in [(7, fmt_kg), (8, fmt_eu), (9, fmt_eu), (10, fmt_qt)]:
        col  = get_column_letter(ci)
        cell = ws.cell(row=TR, column=ci)
        cell.value         = "=SUM(" + col + str(DS) + ":" + col + str(TR - 1) + ")"
        cell.number_format = fmt
        cell.font = tft; cell.fill = tf
        cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[TR].height = 20

    ws.freeze_panes = ws.cell(row=DS, column=1)
    ws.auto_filter.ref = (get_column_letter(1) + str(HR) + ":" +
                          get_column_letter(ncols) + str(TR - 1))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ---------------------------------------------------------------------------
# CSV builder
# ---------------------------------------------------------------------------

def build_csv(nbb_rows):
    buf    = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(HEADERS)
    for row in nbb_rows:
        wt  = str(row["netto_gewicht"]).replace(".", ",") if row["netto_gewicht"] else ""
        val = str(row["waarde"]).replace(".", ",")
        writer.writerow([row["goederencode"], row["gewest"], row["land"], row["aard"],
                         row["incoterm"], row["vervoer"], wt, val, val,
                         str(row["aanv_eenheden"]), row["aanv_eenh"] or ""])
    return buf.getvalue().encode("utf-8-sig")

# ---------------------------------------------------------------------------
# Bestandsnaam
# ---------------------------------------------------------------------------

def _make_base_name(meta, pdf_filename):
    factuurnr = meta.get("factuurnummer", "").strip()
    if factuurnr:
        safe = re.sub(r"[^A-Za-z0-9_\-]", "", factuurnr)
        if safe:
            return safe
    return secure_filename(pdf_filename).rsplit(".", 1)[0] or "intrastat"

# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------

@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    if "pdf" not in request.files:
        return jsonify({"error": "Geen bestand ontvangen."}), 400
    file = request.files["pdf"]
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Upload een geldig PDF-bestand."}), 400
    pdf_bytes = file.read()
    if not pdf_bytes:
        return jsonify({"error": "Het bestand is leeg."}), 400

    try:
        lines, pdf_total, meta = parse_pdf(pdf_bytes)
    except Exception as exc:
        logger.exception("PDF verwerking mislukt")
        return jsonify({"error": "Fout bij verwerken: " + str(exc)}), 500

    if not lines:
        return jsonify({"error": (
            "Geen productlijnen gevonden. "
            "Controleer of de PDF leesbare tekst bevat met een "
            "QUANTITY / VALUE / CUSTOMS-tabelstructuur."
        )}), 422

    aggregated = aggregate_by_cn8(lines)
    nbb_rows   = [build_nbb_row(a) for a in aggregated]
    total_eur  = round(sum(r["waarde"] for r in nbb_rows), 2)

    if pdf_total is not None:
        diff      = round(abs(total_eur - pdf_total), 2)
        check_ok  = diff < 0.02
        check_msg = (
            "Totaal klopt: EUR {:,.2f} = PDF-factuurtotaal".format(total_eur)
            if check_ok else
            "Verschil: berekend EUR {:,.2f} vs PDF-totaal EUR {:,.2f} (diff {:,.2f})".format(
                total_eur, pdf_total, diff)
        )
    else:
        check_ok  = True
        check_msg = "PDF-factuurtotaal niet gevonden (handmatige controle)"

    base      = _make_base_name(meta, file.filename)
    date_str  = datetime.today().strftime("%Y%m%d")
    xlsx_name = "intrastat_" + base + "_" + date_str + ".xlsx"
    csv_name  = "intrastat_" + base + "_" + date_str + ".csv"

    try:
        excel_bytes = build_excel(nbb_rows, meta)
        csv_bytes   = build_csv(nbb_rows)
    except Exception as exc:
        logger.exception("Bestand aanmaken mislukt")
        return jsonify({"error": "Bestand aanmaken mislukt: " + str(exc)}), 500

    token = str(uuid.uuid4())
    with STORE_LOCK:
        RESULT_STORE[token] = {
            "excel": (excel_bytes, xlsx_name),
            "csv":   (csv_bytes,   csv_name),
            "ts":    time.time(),
        }

    return jsonify({
        "token": token, "xlsx_name": xlsx_name, "csv_name": csv_name,
        "item_count": len(nbb_rows), "raw_lines": len(lines),
        "total_eur": total_eur, "pdf_total": pdf_total,
        "check_ok": check_ok, "check_msg": check_msg, "meta": meta,
    })


@app.route("/download/<token>/<fmt>")
@login_required
def download(token, fmt):
    with STORE_LOCK:
        result = RESULT_STORE.get(token)
    if result is None:
        abort(404)
    if fmt not in ("excel", "csv"):
        abort(400)
    data, filename = result[fmt]
    mime = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if fmt == "excel" else "text/csv; charset=utf-8"
    )
    return send_file(io.BytesIO(data), mimetype=mime,
                     as_attachment=True, download_name=filename)


@app.route("/done/<token>")
@login_required
def done(token):
    with STORE_LOCK:
        RESULT_STORE.pop(token, None)
    return "", 204


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
